from flask import request, send_file, jsonify
from io import BytesIO, StringIO
from routes import api
from utils.auth import auth_required
from database import db
from models.product import Product
from models.client import Client
from models.process import Process
from models.provider import Provider
from datetime import datetime
import csv
import re


def _parse_date(s: str | None):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _query_dataset(kind: str):
    if kind == "products":
        return [p.to_dict() for p in db.session.query(Product).all()]
    if kind == "clients":
        return [c.to_dict() for c in db.session.query(Client).all()]
    if kind == "processes":
        q = db.session.query(Process)
        d_from = _parse_date(request.args.get("from"))
        d_to = _parse_date(request.args.get("to"))
        op = request.args.get("op")
        if d_from:
            q = q.filter(Process.created_at >= datetime.combine(d_from, datetime.min.time()))
        if d_to:
            q = q.filter(Process.created_at <= datetime.combine(d_to, datetime.max.time()))
        if op:
            like = f"%{op}%"
            q = q.filter(Process.op.ilike(like))
        return [p.to_dict() for p in q.order_by(Process.created_at.desc()).all()]
    if kind == "providers":
        return [p.to_dict() for p in db.session.query(Provider).all()]
    if kind == "inventory":
        # Build inventory rows from procesos + catalogs (human readable)
        def _ident(name: str) -> str:
            if not name:
                raise ValueError("Empty identifier")
            if not re.fullmatch(r"[A-Za-z0-9_\.]+", name):
                raise ValueError(f"Invalid identifier: {name}")
            return ".".join(f"[{p}]" for p in name.split("."))

        proct = "dbo.procesos"
        cli_tbl = "dbo.clientes"
        prod_tbl = "dbo.productos"
        est_tbl = "dbo.estaciones"
        usu_tbl = "dbo.usuarios"

        def _cols_of(tbl: str) -> set:
            try:
                rs = db.session.execute(
                    db.text("SELECT name FROM sys.columns WHERE object_id = OBJECT_ID(:t)")
                , {"t": tbl}).fetchall()
                return {r[0] for r in rs}
            except Exception:
                return set()

        pcols = _cols_of(proct)
        ccols = _cols_of(cli_tbl)
        has_pr_est = "estacion_id" in pcols
        has_pr_usu = "usuario_id" in pcols
        has_cli_usu = "usuario_id" in ccols
        has_created_at = "created_at" in pcols

        sel = [
            "pr.[id] AS [Folio]",
            "pr.[op] AS [OP]",
            "CASE WHEN c.[nombre] LIKE 'AUTO-%' THEN NULL ELSE c.[nombre] END AS [IdClie]",
            "CASE WHEN p.[nombre] LIKE 'AUTO-%' THEN NULL ELSE p.[nombre] END AS [IdProd]",
            "CASE WHEN pr.[variable1] LIKE 'AUTO-%' THEN NULL ELSE pr.[variable1] END AS [Var1]",
            "CASE WHEN pr.[variable2] LIKE 'AUTO-%' THEN NULL ELSE pr.[variable2] END AS [Var2]",
            "CASE WHEN pr.[variable3] LIKE 'AUTO-%' THEN NULL ELSE pr.[variable3] END AS [Var3]",
            "pr.[piezas] AS [Pzas]",
            "p.[peso_por_pieza] AS [PxP]",
            "ROUND(COALESCE(pr.[piezas],0) * COALESCE(p.[peso_por_pieza],0), 2) AS [Peso]",
            "pr.[lote] AS [Lote]",
            ("CASE WHEN s.[idest] LIKE 'AUTO-%' THEN NULL ELSE s.[idest] END AS [IdEst]" if has_pr_est else "CAST(NULL AS NVARCHAR(50)) AS [IdEst]"),
            ("CASE WHEN u.[nombre] LIKE 'AUTO-%' THEN NULL ELSE u.[nombre] END AS [IdUsu]" if (has_pr_usu or has_cli_usu) else "CAST(NULL AS NVARCHAR(50)) AS [IdUsu]"),
            ("CAST(pr.[created_at] AS DATE) AS [Fecha]" if has_created_at else "CAST(NULL AS DATE) AS [Fecha]"),
        ]

        sql_parts = [
            "SELECT ", ", ".join(sel),
            f" FROM {_ident(proct)} AS pr",
            f" JOIN {_ident(cli_tbl)} AS c ON c.[id] = pr.[cliente_id]",
            f" JOIN {_ident(prod_tbl)} AS p ON p.[id] = pr.[producto_id]",
        ]
        if has_pr_est:
            sql_parts.append(f" LEFT JOIN {_ident(est_tbl)} AS s ON s.[id] = pr.[estacion_id]")
        if has_pr_usu:
            sql_parts.append(f" LEFT JOIN {_ident(usu_tbl)} AS u ON u.[id] = pr.[usuario_id]")
        elif has_cli_usu:
            sql_parts.append(f" LEFT JOIN {_ident(usu_tbl)} AS u ON u.[id] = c.[usuario_id]")

        where = []
        params = {}
        d_from = _parse_date(request.args.get("from"))
        d_to = _parse_date(request.args.get("to"))
        mr = request.args.get("mr")
        if mr:
            where.append("pr.[op] LIKE :mr")
            params["mr"] = f"%{mr}%"
        if d_from and has_created_at:
            where.append("CAST(pr.[created_at] AS DATE) >= :dfrom")
            params["dfrom"] = d_from
        if d_to and has_created_at:
            where.append("CAST(pr.[created_at] AS DATE) <= :dto")
            params["dto"] = d_to
        if where:
            sql_parts.append(" WHERE " + " AND ".join(where))
        sql_parts.append(" ORDER BY pr.[created_at] DESC" if has_created_at else " ORDER BY pr.[id] DESC")
        sql = "".join(sql_parts)
        rows = db.session.execute(db.text(sql), params).mappings().all()
        return [dict(r) for r in rows]
    return None


@api.get("/export/excel")
@auth_required()
def export_excel():
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "Dependencia faltante: openpyxl"}), 500
    kind = request.args.get("kind", "products")
    data = _query_dataset(kind)
    if data is None:
        return jsonify({"error": "Dataset invalido"}), 400
    columns_qs = request.args.get("columns")
    if columns_qs:
        headers = [c for c in columns_qs.split(",") if c]
    else:
        headers = (list(data[0].keys()) if data else [])
    wb = openpyxl.Workbook()
    ws = wb.active
    if data:
        ws.append(headers)
        for row in data:
            ws.append([row.get(k) for k in headers])
        # Autosize
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for r in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                v = r[0].value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(10, int(max_len*1.2)+2))
        ws.freeze_panes = "A2"
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"{kind}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@api.get("/export/pdf")
@auth_required()
def export_pdf():
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, landscape
    except ImportError:
        return jsonify({"error": "Dependencia faltante: reportlab"}), 500
    kind = request.args.get("kind", "products")
    data = _query_dataset(kind)
    if data is None:
        return jsonify({"error": "Dataset invalido"}), 400
    headers = None
    if kind == "inventory":
        columns_qs = request.args.get('columns')
        if columns_qs:
            headers = [c for c in columns_qs.split(',') if c]
    bio = BytesIO()
    c = canvas.Canvas(bio)
    # Page size/orientation
    if kind == "inventory":
        try:
            c.setPageSize(landscape(letter))
        except Exception:
            pass
    # Layout metrics
    width, height = c._pagesize  # current page size
    margin_l = 40
    margin_r = 40
    top_y = height - 40
    table_left = margin_l
    table_right = width - margin_r
    # Header
    c.setFont("Helvetica-Bold", 14)
    y = top_y
    c.drawString(margin_l, y, "Sistema Administrativo - Reporte")
    c.setFont("Helvetica", 10); y -= 16
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    user = getattr(request, 'user', None)
    user_name = getattr(user, 'nombre', 'Usuario') if user else 'Usuario'
    filters = {k: request.args.get(k) for k in ("from","to","mr") if request.args.get(k)}
    c.drawString(margin_l, y, f"Tipo: {kind}  Fecha: {now}  Usuario: {user_name}"); y -= 14
    if filters:
        c.drawString(margin_l, y, f"Filtros: {filters}"); y -= 14
    c.line(margin_l, y, table_right, y); y -= 20
    # Table
    if data:
        if headers is None:
            headers = list(data[0].keys())
        left = table_left; right = table_right
        table_w = right - left
        samples = data[:100]
        lengths = []
        for h in headers:
            m = len(str(h))
            for row in samples:
                v = row.get(h)
                if v is None: continue
                m = max(m, len(str(v)))
            lengths.append(max(6, m))
        total = sum(lengths) or 1
        col_widths = [max(40, int(table_w * (l/total))) for l in lengths]
        row_h = 16
        c.setFillColorRGB(0.12, 0.16, 0.22)
        c.rect(left, y - row_h, table_w, row_h, fill=1, stroke=0)
        c.setFillColorRGB(1, 1, 1); c.setFont("Helvetica-Bold", 9)
        x = left
        for i, h in enumerate(headers):
            c.drawString(x + 2, y - 12, str(h)[:28]); x += col_widths[i]
        y -= (row_h + 4); c.setFont("Helvetica", 9); c.setFillColorRGB(0,0,0)
        for idx, row in enumerate(data):
            if y < 40 + row_h:
                c.showPage()
                # re-apply page size/orientation and metrics
                if kind == "inventory":
                    try:
                        c.setPageSize(landscape(letter))
                    except Exception:
                        pass
                width, height = c._pagesize
                top_y = height - 40
                left = margin_l; right = width - margin_r; table_w = right - left
                y = top_y
                c.setFillColorRGB(0.12, 0.16, 0.22)
                c.rect(left, y - row_h, table_w, row_h, fill=1, stroke=0)
                c.setFillColorRGB(1, 1, 1); c.setFont("Helvetica-Bold", 9)
                x = left
                for i, h in enumerate(headers):
                    c.drawString(x + 2, y - 12, str(h)[:28]); x += col_widths[i]
                y -= (row_h + 4); c.setFont("Helvetica", 9); c.setFillColorRGB(0,0,0)
            if idx % 2 == 0:
                c.setFillColorRGB(0.96, 0.97, 0.99)
                c.rect(left, y - row_h + 2, table_w, row_h, fill=1, stroke=0)
                c.setFillColorRGB(0, 0, 0)
            x = left
            for i, h in enumerate(headers):
                txt = str(row.get(h)); max_chars = max(8, int(col_widths[i] / 6))
                c.drawString(x + 2, y - 12, txt[:max_chars]); x += col_widths[i]
            y -= row_h
    c.save(); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"{kind}.pdf", mimetype="application/pdf")


@api.get("/export/csv")
@auth_required()
def export_csv():
    kind = request.args.get("kind", "products")
    data = _query_dataset(kind)
    if data is None:
        return jsonify({"error": "Dataset invalido"}), 400
    if not data:
        data = [{}]
    headers = list(data[0].keys())
    sio = StringIO(); writer = csv.writer(sio)
    writer.writerow(headers)
    for row in data:
        writer.writerow([row.get(h) for h in headers])
    data_bytes = sio.getvalue().encode("utf-8-sig")
    bio = BytesIO(data_bytes); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"{kind}.csv", mimetype="text/csv; charset=utf-8")
